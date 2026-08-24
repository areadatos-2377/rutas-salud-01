"""
Genera data/clues_catalogo.json a partir de data/raw/CLUES_IMB.xlsx
(catalogo de unidades) y data/raw/ejemplo_6ta_distribucion_BC.xlsx
(hoja Hoja1, catalogo de entidades/coordinadores).

Ejecutar cada vez que se actualice el catalogo de unidades:
    python scripts/generar_catalogo_clues.py
"""

import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
CLUES_XLSX = BASE_DIR / "data" / "raw" / "CLUES_IMB.xlsx"
ENTIDADES_XLSX = BASE_DIR / "data" / "raw" / "ejemplo_6ta_distribucion_BC.xlsx"
OUTPUT_JSON = BASE_DIR / "data" / "clues_catalogo.json"
OUTPUT_JSON_TOOL_COPY = BASE_DIR / "tools" / "captura-programacion" / "clues_catalogo.json"

NIVEL_ATENCION_ESPERADO = "PRIMER NIVEL"
ESTATUS_ESPERADO = "EN OPERACION"

# Entidades donde el nombre en el catalogo de unidades (BD_IMB) difiere
# del nombre oficial usado en el catalogo de entidades/coordinadores (Hoja1).
ALIAS_ENTIDAD = {
    "MEXICO": "ESTADO DE MEXICO",
    "MICHOACAN DE OCAMPO": "MICHOACAN",
    "VERACRUZ DE IGNACIO DE LA LLAVE": "VERACRUZ",
}


def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto


def mapear_tipo_unidad_medica(nombre_tipologia):
    """Deriva el bucket 'TIPO DE UNIDAD MEDICA' visto en el Excel real
    a partir de NOMBRE DE TIPOLOGIA del catalogo. Inferido de una sola
    muestra (528 filas de Baja California) -- validar con el area."""
    t = normalizar(nombre_tipologia)

    if "UNEME" in t or "UNIDAD DE ESPECIALIDADES MEDICAS" in t:
        return "UNEME"
    if "UNIDAD MOVIL" in t:
        return "UNIDAD MOVIL"

    match = re.search(r"(\d{2})\s*NUCLEO", t)
    if match:
        n = int(match.group(1))
        if n <= 2:
            return "1-2 NUCLEOS"
        if n <= 5:
            return "3-5 NUCLEOS"
        return "6-12 NUCLEOS"

    if "01 NUCLEO" in t:
        return "1-2 NUCLEOS"

    # Tipologia sin mapeo conocido (no vista en la muestra real):
    # se conserva el nombre original para no perder informacion.
    return nombre_tipologia


def cargar_coordinadores():
    wb = openpyxl.load_workbook(ENTIDADES_XLSX, data_only=True)
    ws = wb["Hoja1"]
    coordinadores = {}
    orden_entidades = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        entidad, titular = row[0], row[1]
        if not entidad:
            continue
        entidad = str(entidad).strip()
        coordinadores[normalizar(entidad)] = {
            "entidad": entidad,
            "coordinador": (titular or "").strip(),
        }
        orden_entidades.append(entidad)
    return coordinadores, orden_entidades


def cargar_unidades(entidades_validas_norm):
    wb = openpyxl.load_workbook(CLUES_XLSX, data_only=True)
    ws = wb["BD_IMB"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    unidades_por_entidad = {}
    total_filtradas = 0
    total_leidas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_leidas += 1
        nivel = row[idx["NIVEL ATENCION"]]
        estatus = row[idx["ESTATUS DE OPERACION"]]
        if nivel != NIVEL_ATENCION_ESPERADO or estatus != ESTATUS_ESPERADO:
            continue

        entidad_original = row[idx["ENTIDAD"]]
        entidad_norm = normalizar(entidad_original)
        entidad_norm = normalizar(ALIAS_ENTIDAD.get(entidad_norm, entidad_norm))

        if entidad_norm not in entidades_validas_norm:
            continue

        tipologia = row[idx["NOMBRE DE TIPOLOGIA"]]
        unidad = {
            "clues": (row[idx["CLUES"]] or "").strip(),
            "nombre_unidad": (row[idx["NOMBRE DE LA UNIDAD"]] or "").strip(),
            "tipo_unidad_medica": mapear_tipo_unidad_medica(tipologia),
            "tipologia_original": tipologia,
            "municipio": row[idx["MUNICIPIO"]],
        }
        if not unidad["clues"]:
            continue

        unidades_por_entidad.setdefault(entidad_norm, []).append(unidad)
        total_filtradas += 1

    print(f"Filas leidas en BD_IMB: {total_leidas}")
    print(f"Filas incluidas (primer nivel, en operacion, entidad reconocida): {total_filtradas}")
    return unidades_por_entidad


def main():
    coordinadores, orden_entidades = cargar_coordinadores()
    entidades_validas_norm = set(coordinadores.keys())
    unidades_por_entidad = cargar_unidades(entidades_validas_norm)

    salida = {
        "generado_en": datetime.now(timezone.utc).isoformat(),
        "fuente_unidades": str(CLUES_XLSX.relative_to(BASE_DIR)).replace("\\", "/"),
        "fuente_entidades": str(ENTIDADES_XLSX.relative_to(BASE_DIR)).replace("\\", "/"),
        "entidades": {},
    }

    for entidad in orden_entidades:
        entidad_norm = normalizar(entidad)
        info_coord = coordinadores[entidad_norm]
        unidades = sorted(
            unidades_por_entidad.get(entidad_norm, []),
            key=lambda u: u["nombre_unidad"],
        )
        salida["entidades"][entidad] = {
            "coordinador_sugerido": info_coord["coordinador"],
            "total_unidades": len(unidades),
            "unidades": unidades,
        }
        print(f"  {entidad}: {len(unidades)} unidades")

    for destino in (OUTPUT_JSON, OUTPUT_JSON_TOOL_COPY):
        destino.parent.mkdir(parents=True, exist_ok=True)
        with open(destino, "w", encoding="utf-8") as f:
            json.dump(salida, f, ensure_ascii=False, indent=2)
        print(f"Escrito: {destino.relative_to(BASE_DIR)}")


if __name__ == "__main__":
    main()

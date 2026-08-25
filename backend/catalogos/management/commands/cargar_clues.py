"""
Carga/actualiza el catalogo de Entidad y UnidadMedica en la base de datos
real desde data/raw/CLUES_IMB.xlsx y data/raw/ejemplo_6ta_distribucion_BC.xlsx
(este ultimo solo para el catalogo de entidades/coordinadores, hoja Hoja1).

Reusa la misma logica de lectura/mapeo que scripts/generar_catalogo_clues.py
(ese genera el JSON que consume la sub-herramienta estatica; este comando
escribe directo en la base de datos del backend -- son consumidores
distintos de la misma fuente, por eso no comparten codigo entre un
entorno Python global y el venv de Django).

Seguro de re-ejecutar cada vez que se reemplace el Excel (actualizacion
mensual, segun el area): hace upsert por CLUES -- agrega los nuevos,
actualiza los que cambiaron de nombre/tipo/municipio, y NUNCA toca ni
borra las unidades con origen=manual (las que alguien capturo a mano
porque su CLUES no estaba en el catalogo oficial). Tampoco borra CLUES
que hayan desaparecido del nuevo Excel -- si algun dia hace falta dar de
baja unidades formalmente, es una decision de negocio aparte, no algo
que este comando deba hacer solo.
"""

import re
import unicodedata
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from catalogos.models import Entidad, UnidadMedica

BASE_DIR = Path(__file__).resolve().parents[4]
CLUES_XLSX = BASE_DIR / "data" / "raw" / "CLUES_IMB.xlsx"
ENTIDADES_XLSX = BASE_DIR / "data" / "raw" / "ejemplo_6ta_distribucion_BC.xlsx"

# Decision 2026-08-24: se amplia de solo primer nivel (blueprint-v00) a los
# 3 niveles reales de la fuente -- excluye "NO APLICA", que en CLUES_IMB.xlsx
# corresponde a registros administrativos, no unidades de atencion.
NIVELES_ATENCION_ESPERADOS = {
    UnidadMedica.NIVEL_PRIMER,
    UnidadMedica.NIVEL_SEGUNDO,
    UnidadMedica.NIVEL_TERCER,
}
ESTATUS_ESPERADO = "EN OPERACION"

ALIAS_ENTIDAD = {
    "MEXICO": "ESTADO DE MEXICO",
    "MICHOACAN DE OCAMPO": "MICHOACAN",
    "VERACRUZ DE IGNACIO DE LA LLAVE": "VERACRUZ",
}


def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto)


def mapear_tipo_unidad_medica(nombre_tipologia):
    t = normalizar(nombre_tipologia)
    if "UNEME" in t or "UNIDAD DE ESPECIALIDADES MEDICAS" in t:
        return "UNEME"
    if "UNIDAD MOVIL" in t:
        return "UNIDAD MOVIL"
    match = re.search(r"(\d{2})\s*NUCLEO", t)
    if match:
        n = int(match.group(1))
        if n <= 2:
            return "1-2 NUCLEOS"
        if n <= 5:
            return "3-5 NUCLEOS"
        return "6-12 NUCLEOS"
    if "01 NUCLEO" in t:
        return "1-2 NUCLEOS"
    return nombre_tipologia


class DryRunRollback(Exception):
    pass


class Command(BaseCommand):
    help = (
        "Carga/actualiza Entidad y UnidadMedica desde data/raw/CLUES_IMB.xlsx. "
        "Re-ejecutar cada vez que se reemplace ese Excel (actualizacion mensual)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Muestra que haria sin escribir en la base de datos.",
        )

    def handle(self, *args, **options):
        if not CLUES_XLSX.exists():
            raise CommandError(f"No se encontro {CLUES_XLSX}")
        if not ENTIDADES_XLSX.exists():
            raise CommandError(f"No se encontro {ENTIDADES_XLSX}")

        dry_run = options["dry_run"]

        try:
            with transaction.atomic():
                self._cargar(dry_run)
                if dry_run:
                    raise DryRunRollback()
        except DryRunRollback:
            self.stdout.write(self.style.WARNING("\n--dry-run: no se escribio nada en la base de datos."))

    def _cargar(self, dry_run):
        coordinadores, orden_entidades = self._leer_coordinadores()

        mapa_entidad = {}
        entidades_creadas = entidades_actualizadas = 0
        for nombre in orden_entidades:
            coordinador = coordinadores[normalizar(nombre)]
            entidad, creada = Entidad.objects.update_or_create(
                nombre=nombre, defaults={"coordinador": coordinador}
            )
            mapa_entidad[normalizar(nombre)] = entidad
            entidades_creadas += creada
            entidades_actualizadas += not creada

        self.stdout.write(
            f"Entidades: {entidades_creadas} creadas, {entidades_actualizadas} actualizadas "
            f"(de {len(orden_entidades)} en el catalogo)."
        )

        creadas = actualizadas = sin_cambios = 0
        omitidas_entidad_no_reconocida = 0
        total_leidas = 0

        wb = openpyxl.load_workbook(CLUES_XLSX, data_only=True)
        ws = wb["BD_IMB"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            total_leidas += 1
            nivel_atencion = row[idx["NIVEL ATENCION"]]
            if nivel_atencion not in NIVELES_ATENCION_ESPERADOS:
                continue
            if row[idx["ESTATUS DE OPERACION"]] != ESTATUS_ESPERADO:
                continue

            entidad_norm = normalizar(row[idx["ENTIDAD"]])
            entidad_norm = normalizar(ALIAS_ENTIDAD.get(entidad_norm, entidad_norm))
            entidad_obj = mapa_entidad.get(entidad_norm)
            if entidad_obj is None:
                omitidas_entidad_no_reconocida += 1
                continue

            clues = (row[idx["CLUES"]] or "").strip()
            if not clues:
                continue

            tipologia = row[idx["NOMBRE DE TIPOLOGIA"]]
            defaults = {
                "nombre": (row[idx["NOMBRE DE LA UNIDAD"]] or "").strip(),
                "entidad": entidad_obj,
                "tipo_unidad_medica": mapear_tipo_unidad_medica(tipologia),
                "municipio": row[idx["MUNICIPIO"]] or "",
                "origen": UnidadMedica.ORIGEN_CATALOGO_MENSUAL,
                "nivel_atencion": nivel_atencion,
            }

            anterior = UnidadMedica.objects.filter(pk=clues).first()
            _, creada = UnidadMedica.objects.update_or_create(clues=clues, defaults=defaults)
            if creada:
                creadas += 1
            elif anterior and all(getattr(anterior, k) == v for k, v in defaults.items() if k != "entidad"):
                sin_cambios += 1
            else:
                actualizadas += 1

        self.stdout.write(
            f"Unidades medicas: {creadas} creadas, {actualizadas} actualizadas, "
            f"{sin_cambios} sin cambios (de {total_leidas} filas leidas en el Excel)."
        )
        if omitidas_entidad_no_reconocida:
            self.stdout.write(
                self.style.WARNING(
                    f"{omitidas_entidad_no_reconocida} filas omitidas: entidad no reconocida "
                    "(fuera de las 23 entidades del catalogo, ej. Guanajuato/Yucatan)."
                )
            )
        self.stdout.write(
            self.style.SUCCESS(
                f"Total en catalogo tras la carga: {UnidadMedica.objects.filter(origen=UnidadMedica.ORIGEN_CATALOGO_MENSUAL).count()} "
                f"unidades de catalogo mensual + {UnidadMedica.objects.filter(origen=UnidadMedica.ORIGEN_MANUAL).count()} capturadas manualmente."
            )
        )

    def _leer_coordinadores(self):
        wb = openpyxl.load_workbook(ENTIDADES_XLSX, data_only=True)
        ws = wb["Hoja1"]
        coordinadores = {}
        orden = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            entidad, titular = row[0], row[1]
            if not entidad:
                continue
            entidad = str(entidad).strip()
            coordinadores[normalizar(entidad)] = (titular or "").strip()
            orden.append(entidad)
        return coordinadores, orden
